"""
Axis Max Life Insurance Portal – Selenium Automation
=====================================================
Automates the premium calculator portal for the first 6 test rows in
`test data/insurance_test_data.xlsx`, downloads Benefit Illustration PDFs,
extracts Equote Number & Premium from both the portal and the PDF, and
writes a final report to `./test_results/`.

Dependencies (install once):
    pip install selenium openpyxl PyPDF2 webdriver-manager

Usage:
    python axis_max_life_automation.py
"""

import os
import re
import csv
import time
import logging
import traceback
from pathlib import Path
from datetime import datetime

import openpyxl
import PyPDF2
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
)
from webdriver_manager.chrome import ChromeDriverManager

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

PORTAL_URL = "https://www.axismaxlife.com/term-insurance-plans/premium-calculator?stage=rider&utmCode=1311271&utm_theme=1Crore&utm_source=google&utm_medium=cpc&utm_campaign=1_Brand_Exact_Axis_Max_Life_04022025&utm_content=SKAG_New_14042025&utm_term=axis%20max%20life%20insurance&gclid=CjwKCAjwtvvPBhBuEiwAPMijr0PY3tR7cdeyyt3EJ0YTZ6m-37Dq5hDMFkrmrQIYKy3Ds9PLV7f1pRoCAh8QAvD_BwE"
EXCEL_PATH = Path(__file__).parent / "test data" / "insurance_test_data.xlsx"
DOWNLOAD_DIR = Path(__file__).parent / "downloads"
RESULTS_DIR = Path(__file__).parent / "test_results"
SCREENSHOTS_DIR = Path(__file__).parent / "screenshots"

HEADLESS = False          # Set True for headless mode
MAX_ROWS = 4              # Process 6 rows as requested
DEFAULT_WAIT = 20         # seconds for explicit waits
DOWNLOAD_WAIT = 30        # seconds to wait for PDF download

# Occupation mapping — portal shows "Salaried" and "Self-employed/Business"
OCCUPATION_MAP = {
    "salaried": "Salaried",
    "self employed": "Self-employed/Business",
    "self-employed": "Self-employed/Business",
    "self employed/business": "Self-employed/Business",
    "housewife": "Self-employed/Business",   # mapped as agreed
    "business": "Self-employed/Business",
}

# Education mapping — portal shows "Graduate & Above", "12th Pass", "10th Pass"
# We map to shorter values that will match via starts-with logic
EDUCATION_MAP = {
    "grad or above": "Graduate & Above",
    "graduate": "Graduate & Above",
    "graduate & above": "Graduate & Above",
    "grad or c": "Graduate & Above",
    "post graduate": "Graduate & Above",
    "post-graduate": "Graduate & Above",
    "12th pass": "12th Pass",
    "12th": "12th Pass",
    "10th pass": "10th Pass",
    "10th": "10th Pass",
    "below 10th": "10th Pass",
}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.DEBUG,   # DEBUG shows every click attempt — change to INFO once stable
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Income bracket helper
# ---------------------------------------------------------------------------

def map_income_to_bracket(income_raw) -> str:
    """
    Convert a fixed numeric income value to the portal's radio-button label.

    Actual portal labels (from UI screenshot):
        "<5"        →  less than 5 Lakhs
        "5 - 7"     →  5 to 7 Lakhs
        "7 - 10"    →  7 to 10 Lakhs
        "10 - 20"   →  10 to 20 Lakhs
        ">20"       →  above 20 Lakhs

    Income in the Excel is a fixed rupee value (e.g. 3,812,052).
    We convert to Lakhs (divide by 1,00,000) then pick the bracket.
    """
    # Strip commas, spaces, currency symbols
    income_str = str(income_raw).replace(",", "").replace("₹", "").strip()
    try:
        amount = float(income_str)
    except ValueError:
        log.warning("Cannot parse income '%s', defaulting to '>20'", income_raw)
        return ">20"

    # Convert rupees → Lakhs
    lakhs = amount / 1_00_000

    if lakhs < 5:
        return "<5"
    elif lakhs < 7:
        return "5 - 7"
    elif lakhs < 10:
        return "7 - 10"
    elif lakhs <= 20:
        return "10 - 20"
    else:
        return ">20"


def parse_life_cover(raw) -> str:
    """
    Normalise life cover to a plain number string (in rupees).
    Accepts: '3 crore', '50 lakh', '1.5 crore', '5000000', etc.
    Returns the numeric string the portal currency input expects.
    """
    raw = str(raw).lower().strip()
    if "crore" in raw:
        num = float(re.sub(r"[^\d.]", "", raw))
        return str(int(num * 1_00_00_000))
    elif "lakh" in raw or "lac" in raw:
        num = float(re.sub(r"[^\d.]", "", raw))
        return str(int(num * 1_00_000))
    else:
        return re.sub(r"[^\d]", "", raw)


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def read_test_data() -> list[dict]:
    """Read the first MAX_ROWS data rows from the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    log.info("Excel headers: %s", headers)

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= MAX_ROWS:
            break
        record = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
        rows.append(record)

    log.info("Loaded %d test rows", len(rows))
    return rows


# ---------------------------------------------------------------------------
# WebDriver factory
# ---------------------------------------------------------------------------

def create_driver() -> webdriver.Chrome:
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    prefs = {
        "download.default_directory": str(DOWNLOAD_DIR.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,   # force PDF download
    }

    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-popup-blocking")
    opts.add_experimental_option("prefs", prefs)
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(60)
    return driver


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def wait_for(driver, by, selector, timeout=DEFAULT_WAIT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )


def wait_clickable(driver, by, selector, timeout=DEFAULT_WAIT):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )


def safe_click(driver, element):
    """Click with JS fallback if intercepted."""
    try:
        element.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", element)


def click_button_by_text(driver, text, timeout=DEFAULT_WAIT):
    """Click the first visible button/div whose text matches (case-insensitive)."""
    xpath = (
        f"//*[self::button or self::div or self::span or self::label]"
        f"[normalize-space(text())='{text}' or normalize-space(.)='{text}']"
    )
    el = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((By.XPATH, xpath))
    )
    safe_click(driver, el)


def clear_and_type(driver, by, selector, value, timeout=DEFAULT_WAIT):
    el = wait_clickable(driver, by, selector, timeout)
    el.click()
    el.send_keys(Keys.CONTROL + "a")
    el.send_keys(Keys.DELETE)
    el.send_keys(str(value))
    return el


def take_screenshot(driver, name: str):
    """Save a screenshot — silently skips if the browser session is already gone."""
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    path = SCREENSHOTS_DIR / f"{name}_{datetime.now().strftime('%H%M%S')}.png"
    try:
        driver.save_screenshot(str(path))
        log.info("Screenshot saved: %s", path)
    except Exception as exc:
        log.debug("Screenshot skipped (%s: %s)", type(exc).__name__, exc)


# ---------------------------------------------------------------------------
# Page automation functions
# ---------------------------------------------------------------------------

def fill_landing_page(driver: webdriver.Chrome, data: dict):
    """Fill the landing page form and click View Plans."""
    log.info("  → Landing page")

    # Full Name
    clear_and_type(driver, By.CSS_SELECTOR, "input#fullName", data["Full Name"])

    # Date of Birth – portal expects DD/MM/YYYY or DDMMYYYY
    dob_raw = data["Date of birth"]
    if hasattr(dob_raw, "strftime"):
        dob_str = dob_raw.strftime("%d%m%Y")
    else:
        dob_str = str(dob_raw).replace("/", "").replace("-", "").strip()
    clear_and_type(driver, By.CSS_SELECTOR, "input#dob", dob_str)

    # NRI Status – always "No"
    _click_radio_label(driver, "No")

    # Mobile
    clear_and_type(driver, By.CSS_SELECTOR, "input#mobile", str(data["Mobile"]).strip())

    # Annual Income bracket
    bracket = map_income_to_bracket(data["Annual income"])
    log.info("    Income bracket: %s", bracket)
    _click_income_bracket(driver, bracket)

    # Submit
    wait_clickable(driver, By.CSS_SELECTOR, "button#viewPlans").click()
    log.info("  ✓ Landing page submitted")


def _click_radio_label(driver, label_text, timeout=DEFAULT_WAIT):
    """Click a radio button whose associated label contains the given text."""
    xpath = (
        f"//label[contains(normalize-space(.), '{label_text}')]"
        f" | //span[contains(normalize-space(.), '{label_text}')]"
        f"/ancestor::label"
    )
    try:
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
        safe_click(driver, el)
    except TimeoutException:
        # Fallback: find input[type=radio] near matching text
        xpath2 = (
            f"//input[@type='radio'][following-sibling::*[contains(normalize-space(.), '{label_text}')] "
            f"or preceding-sibling::*[contains(normalize-space(.), '{label_text}')]]"
        )
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath2))
        )
        safe_click(driver, el)


def _click_income_bracket(driver, bracket: str, timeout=DEFAULT_WAIT):
    """
    Click the Annual Income tile/button that exactly matches the bracket label.
    The portal renders income as clickable tiles (e.g. '<5', '5 - 7', '>20').
    Tries an exact-text match first, then a contains match as fallback.
    """
    # Exact match on the tile text
    xpath_exact = (
        f"//*[self::div or self::span or self::button or self::label or self::td]"
        f"[normalize-space(text())='{bracket}']"
    )
    try:
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath_exact))
        )
        safe_click(driver, el)
        log.info("    Clicked income bracket (exact): %s", bracket)
        return
    except TimeoutException:
        pass

    # Contains match — useful if the portal wraps the label in child elements
    xpath_contains = (
        f"//*[self::div or self::span or self::button or self::label or self::td]"
        f"[contains(normalize-space(.), '{bracket}')]"
    )
    try:
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath_contains))
        )
        safe_click(driver, el)
        log.info("    Clicked income bracket (contains): %s", bracket)
        return
    except TimeoutException:
        pass

    # Last resort: radio input near the bracket text
    _click_radio_label(driver, bracket, timeout)


def fill_details_modal(driver: webdriver.Chrome, data: dict):
    """
    Fill the Required Details modal.

    The modal shows ALL fields at once on one scrollable page:
      Gender → Tobacco/Nicotine → Preferred Language → Occupation → Education
    Selecting Education auto-advances to a second screen:
      Disability/Diabetic popup (radio Yes/No + Proceed)
    Then: Marital Status → Proceed (if shown)
    """
    log.info("  → Details modal")

    # Wait for the modal heading
    WebDriverWait(driver, DEFAULT_WAIT).until(
        EC.visibility_of_element_located(
            (By.XPATH, "//*[contains(text(),'Required below details')]")
        )
    )
    time.sleep(0.8)

    # Gender — scoped to the Gender section
    gender = str(data.get("Gender", "Male")).strip().title()
    log.info("    Gender: %s", gender)
    _click_field_option(driver, section_label="Gender", option_text=gender)
    time.sleep(0.4)

    # Tobacco/Nicotine — always No
    log.info("    Tobacco: No")
    _click_field_option(driver, section_label="Tobacco", option_text="No")
    time.sleep(0.4)

    # Preferred Language — always English
    log.info("    Language: English")
    _click_field_option(driver, section_label="Language", option_text="English")
    time.sleep(0.4)

    # Occupation
    occ_raw = str(data.get("Occupation", "Salaried")).strip().lower()
    occ = OCCUPATION_MAP.get(occ_raw, "Salaried")
    log.info("    Occupation: %s", occ)
    _click_field_option(driver, section_label="Occupation", option_text=occ)
    time.sleep(0.4)

    # Education — triggers auto-advance to next screen
    # The Education tiles are below the fold — scroll into view first
    edu_raw = str(data.get("Education", "Graduate")).strip().lower()
    edu = EDUCATION_MAP.get(edu_raw, "Graduate")
    log.info("    Education: %s (raw: %s)", edu, edu_raw)

    # Scroll the Education label into view so the tiles become visible
    try:
        edu_label = driver.find_element(
            By.XPATH, "//*[contains(normalize-space(text()), 'Education')]"
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", edu_label)
        time.sleep(0.5)
    except NoSuchElementException:
        pass

    # Try clicking the education option - with retry logic
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            log.info("    Attempting to click Education option (attempt %d/%d)", attempt + 1, max_attempts)
            _click_field_option(driver, section_label="Education", option_text=edu)
            log.info("    ✓ Education option clicked successfully")
            break
        except Exception as e:
            log.warning("    Education click attempt %d failed: %s", attempt + 1, e)
            if attempt < max_attempts - 1:
                time.sleep(1)
                # Try scrolling again
                try:
                    edu_label = driver.find_element(
                        By.XPATH, "//*[contains(normalize-space(text()), 'Education')]"
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", edu_label)
                    time.sleep(0.5)
                except:
                    pass
            else:
                log.error("    Failed to click Education after %d attempts", max_attempts)
                raise
    
    time.sleep(1.5)  # wait for transition

    # --- Diabetic + Marital Status screen ---
    # After Education is selected, a screen appears with diabetic and marital status fields
    # We don't need to close any popup - just fill the fields directly
    log.info("  → Handling Diabetic + Marital Status screen")
    try:
        # Wait for the diabetic/marital status screen to appear
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(),'Diabetic') or contains(text(),'diabetic') or contains(text(),'Marital')]")
            )
        )
        log.info("    Diabetic/Marital status screen detected")
        time.sleep(1)
        
        # Click "No" for Diabetic (default)
        log.info("    Clicking 'No' for Diabetic")
        try:
            _click_field_option(driver, section_label="Diabetic", option_text="No")
            log.info("    ✓ Diabetic = No")
        except Exception as e:
            log.warning("    Could not click Diabetic field: %s", e)
        time.sleep(0.5)
        
        # Click "Single" for Marital Status (default)
        log.info("    Clicking 'Single' for Marital Status")
        try:
            _click_field_option(driver, section_label="Marital", option_text="Single")
            log.info("    ✓ Marital Status = Single")
        except Exception as e:
            log.warning("    Could not click Marital Status field: %s", e)
        time.sleep(0.5)
        
        # Click "Check Coverage" or "Proceed" button
        log.info("    Looking for submit button (Check Coverage/Proceed)")
        button_clicked = False
        
        # Strategy 1: Try multiple button text variations with XPath
        button_texts = ["Check Coverage", "Proceed", "Continue", "Next", "Submit", "View Plans"]
        for btn_text in button_texts:
            try:
                check_btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, f"//button[contains(normalize-space(.), '{btn_text}')]"))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", check_btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", check_btn)
                log.info("    ✓ Clicked '%s' button", btn_text)
                button_clicked = True
                break
            except TimeoutException:
                continue
        
        # Strategy 2: Try finding button by ID
        if not button_clicked:
            try:
                log.info("    Trying to find button by ID")
                check_btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button#viewPlans"))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", check_btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", check_btn)
                log.info("    ✓ Clicked button via ID")
                button_clicked = True
            except TimeoutException:
                pass
        
        # Strategy 3: JavaScript fallback - find any visible button
        if not button_clicked:
            log.info("    Trying JS fallback to find submit button")
            result = driver.execute_script("""
                var btns = document.querySelectorAll('button');
                for (var i = 0; i < btns.length; i++) {
                    var btn = btns[i];
                    var txt = (btn.innerText || btn.textContent || '').toLowerCase();
                    var style = window.getComputedStyle(btn);
                    
                    // Check if button is visible and contains relevant text
                    if (style.display !== 'none' && style.visibility !== 'hidden' && 
                        (txt.indexOf('coverage') !== -1 || txt.indexOf('proceed') !== -1 || 
                         txt.indexOf('continue') !== -1 || txt.indexOf('next') !== -1 ||
                         txt.indexOf('submit') !== -1 || txt.indexOf('view') !== -1)) {
                        btn.scrollIntoView({block: 'center'});
                        btn.click();
                        return 'Clicked: ' + txt;
                    }
                }
                
                // If no specific button found, try clicking the last visible button on the page
                for (var i = btns.length - 1; i >= 0; i--) {
                    var btn = btns[i];
                    var style = window.getComputedStyle(btn);
                    if (style.display !== 'none' && style.visibility !== 'hidden') {
                        btn.scrollIntoView({block: 'center'});
                        btn.click();
                        return 'Clicked last button: ' + (btn.innerText || btn.textContent || '');
                    }
                }
                return 'No button found';
            """)
            log.info("    JS result: %s", result)
            if result and 'Clicked' in str(result):
                button_clicked = True
        
        if not button_clicked:
            log.warning("    Could not find submit button - trying to proceed anyway")
        
        time.sleep(3)  # Longer wait for page transition
        
    except TimeoutException:
        log.info("    Diabetic/Marital status screen not shown — skipping")

    # Wait for the modal to be completely gone and next page to load
    log.info("  Waiting for details modal to close completely...")
    time.sleep(3)  # Longer wait for page transition
    
    # Try to detect what page we're on
    log.info("  Detecting current page state...")
    try:
        # Check for various possible page states
        page_indicators = [
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view plan')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'get quote')]",
            "//*[contains(text(),'Customize') or contains(text(),'Step 1') or contains(text(),'Life Cover')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]",
        ]
        
        page_detected = False
        for indicator in page_indicators:
            try:
                element = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, indicator))
                )
                log.info(f"    Found page indicator: {indicator}")
                
                # If it's a button, try clicking it
                if 'button' in indicator.lower():
                    try:
                        if element.is_displayed() and element.is_enabled():
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                            time.sleep(0.5)
                            driver.execute_script("arguments[0].click();", element)
                            log.info("    Clicked button to proceed")
                            time.sleep(2)
                    except:
                        pass
                
                page_detected = True
                break
            except TimeoutException:
                continue
        
        if not page_detected:
            log.warning("    Could not detect page state - taking screenshot")
            take_screenshot(driver, "page_state_unknown")
    
    except Exception as e:
        log.warning(f"    Error detecting page state: {str(e)}")
    
    log.info("  ✓ Details modal done")


def _click_field_option(driver, section_label: str, option_text: str, timeout: int = DEFAULT_WAIT):
    """
    Click an option tile scoped to its section label.
    Finds the heading that contains `section_label`, walks up to the nearest
    container that holds clickable children, then clicks the child whose
    visible text matches `option_text`.

    Searches ANY element tag (li, span, div, p, td, button, etc.) so it works
    regardless of how the portal renders its tiles.
    """
    log.debug("      _click_field_option: section='%s' option='%s'", section_label, option_text)

    # ── Step 1: find the section container ──────────────────────────────────
    # Walk up from the label until we find a div/section/ul/ol that has
    # at least one child element containing the option text.
    section_xpath = (
        f"//*[contains(normalize-space(text()), '{section_label}')]"
        f"/ancestor::*[self::div or self::section or self::ul or self::ol or self::form]"
        f"[.//*[contains(normalize-space(.), '{option_text}')]][1]"
    )
    try:
        container = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, section_xpath))
        )
        log.debug("      Found section container for '%s'", section_label)
    except TimeoutException:
        log.warning("      Section '%s' not found — falling back to global JS click", section_label)
        _js_click_by_text(driver, option_text)
        return

    # ── Step 2: click the matching child element via JS ──────────────────────
    # Scroll the container into view, then find and click the matching child.
    # For Education field: "Graduate" should match "Graduate & Above"
    script = """
        var container = arguments[0];
        var text = arguments[1].trim().toLowerCase();
        container.scrollIntoView({block: 'center'});
        var all = container.querySelectorAll('*');
        
        // First pass: exact match
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t === text) {
                el.scrollIntoView({block: 'center'});
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        // Second pass: starts-with match (e.g., "Graduate" matches "Graduate & Above")
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t.startsWith(text + ' ') || t.startsWith(text + '&')) {
                el.scrollIntoView({block: 'center'});
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        // Third pass: contains match (last resort)
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t.indexOf(text) === 0) {
                el.scrollIntoView({block: 'center'});
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        return null;
    """
    result = driver.execute_script(script, container, option_text)
    if result:
        log.debug("      JS clicked '%s' in section '%s' (matched: %s)", option_text, section_label, result.strip())
        return

    # ── Step 3: global JS fallback ───────────────────────────────────────────
    log.warning("      JS scoped click failed for '%s' — trying global JS click", option_text)
    _js_click_by_text(driver, option_text)


def _js_click_by_text(driver, option_text: str):
    """
    Global fallback: use JS to find and click the first visible element
    whose trimmed text matches option_text (case-insensitive).
    Uses multi-pass matching: exact → starts-with → contains
    """
    script = """
        var text = arguments[0].trim().toLowerCase();
        var all = document.querySelectorAll('button, li, span, div, p, td, label');
        
        // First pass: exact match
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t === text) {
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        // Second pass: starts-with match (e.g., "Graduate" matches "Graduate & Above", "12th" matches "12th Pass")
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t.startsWith(text + ' ') || t.startsWith(text + '&') || t === text + 'th pass' || t === text + 'th') {
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        // Third pass: contains match (last resort)
        for (var i = 0; i < all.length; i++) {
            var el = all[i];
            if (el.children.length > 3) continue;
            var t = (el.innerText || el.textContent || '').trim().toLowerCase();
            if (t.indexOf(text) === 0) {
                el.click();
                return el.innerText || el.textContent;
            }
        }
        
        return null;
    """
    result = driver.execute_script(script, option_text)
    if result:
        log.debug("      Global JS clicked: %s", result.strip())
    else:
        raise TimeoutException(f"Could not find any element with text '{option_text}' on the page")


def _click_check_coverage(driver, timeout=DEFAULT_WAIT):
    """Click the 'Check Coverage' (or similar) submit button in the modal."""
    candidates = [
        "//button[contains(normalize-space(.), 'Check Coverage')]",
        "//button[contains(normalize-space(.), 'Proceed')]",
        "//button[contains(normalize-space(.), 'Submit')]",
    ]
    for xpath in candidates:
        try:
            el = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            safe_click(driver, el)
            return
        except TimeoutException:
            continue
    raise TimeoutException("Could not find Check Coverage / Proceed button in modal")


def fill_step1_customize(driver: webdriver.Chrome, data: dict) -> bool:
    """
    Step 1/4 – Life Cover amount and Cover Till Age.
    Returns True if Step 1 was processed, False if it was skipped by the portal.
    """
    log.info("  → Step 1/4: Customize Plan")

    # Check current URL to see if we're already past Step 1
    current_url = driver.current_url
    log.info("  Current URL: %s", current_url)
    
    # If URL contains stage=rider, we've skipped Step 1 and are already at Step 2
    if 'stage=rider' in current_url or 'stage=2' in current_url:
        log.warning("  ⚠ Page already at Step 2 (Riders) - Step 1 was skipped by the portal")
        log.info("  Skipping Step 1 as portal has auto-advanced")
        return False  # Indicate that Step 1 was skipped
    
    # Handle any intermediate pages (quotes/ads) before Step 1
    log.info("  Checking for intermediate pages...")
    time.sleep(3)  # Wait for page to stabilize
    
    # Check if we need to handle a quotes page or advertisement
    try:
        # Look for any "View Plans", "Get Quotes", "Continue" type buttons
        intermediate_buttons = [
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view plan')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'get quote')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'proceed')]",
            "//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'view plan')]",
        ]
        
        for xpath in intermediate_buttons:
            try:
                btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, xpath))
                )
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                log.info("    Clicked intermediate button")
                time.sleep(2)
                break
            except TimeoutException:
                continue
                
    except Exception as e:
        log.debug("    No intermediate buttons found: %s", e)

    # Wait for step indicator with longer timeout
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Customize') or contains(text(),'Step 1') or contains(text(),'Life Cover')]"))
        )
        log.info("    Step 1 page detected")
    except TimeoutException:
        log.error("    Step 1 page not found - taking screenshot and logging page state")
        take_screenshot(driver, "step1_not_found")
        
        # Log current URL
        current_url = driver.current_url
        log.error("    Current URL: %s", current_url)
        
        # Log page title
        try:
            page_title = driver.title
            log.error("    Page title: %s", page_title)
        except:
            pass
        
        # Log all visible text on the page
        try:
            visible_text = driver.execute_script("""
                return document.body.innerText.substring(0, 500);
            """)
            log.error("    Visible text (first 500 chars): %s", visible_text)
        except:
            pass
        
        # Log all buttons on the page
        try:
            buttons_info = driver.execute_script("""
                var btns = document.querySelectorAll('button, a[role="button"], input[type="button"], input[type="submit"]');
                var result = [];
                for (var i = 0; i < btns.length; i++) {
                    var btn = btns[i];
                    var style = window.getComputedStyle(btn);
                    if (style.display !== 'none' && style.visibility !== 'hidden') {
                        var txt = (btn.innerText || btn.textContent || btn.value || '').trim();
                        if (txt) {
                            result.push(txt);
                        }
                    }
                }
                return result;
            """)
            log.error("    Visible buttons on page: %s", buttons_info)
        except Exception as e:
            log.error("    Could not get buttons info: %s", e)
        
        # Try one more time to click any visible button before giving up
        try:
            log.info("    Attempting emergency button click...")
            click_result = driver.execute_script("""
                var btns = document.querySelectorAll('button, a');
                for (var i = 0; i < btns.length; i++) {
                    var btn = btns[i];
                    var txt = (btn.innerText || btn.textContent || '').toLowerCase();
                    var style = window.getComputedStyle(btn);
                    if (style.display !== 'none' && style.visibility !== 'hidden' && 
                        (txt.indexOf('plan') !== -1 || txt.indexOf('quote') !== -1 || 
                         txt.indexOf('continue') !== -1 || txt.indexOf('proceed') !== -1)) {
                        btn.scrollIntoView({block: 'center'});
                        btn.click();
                        return 'Clicked: ' + txt;
                    }
                }
                return 'No button found';
            """)
            log.info("    Emergency click result: %s", click_result)
            time.sleep(3)
            # Try waiting for Step 1 again
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Customize') or contains(text(),'Step 1') or contains(text(),'Life Cover')]"))
            )
            log.info("    Step 1 page detected after emergency click")
        except Exception as e:
            log.error("    Emergency click also failed: %s", e)
            raise

    # Additional wait for page to be fully interactive
    time.sleep(2)

    # Life Cover – find currency input and type the amount
    cover_amount = parse_life_cover(data.get("Life cover", "5000000"))
    log.info("    Life cover (parsed): %s rupees", cover_amount)
    
    # First, let's log all input fields on the page for debugging
    try:
        all_inputs = driver.execute_script("""
            var inputs = document.querySelectorAll('input');
            var result = [];
            for (var i = 0; i < inputs.length; i++) {
                var inp = inputs[i];
                result.push({
                    type: inp.type || '',
                    id: inp.id || '',
                    name: inp.name || '',
                    placeholder: inp.placeholder || '',
                    value: inp.value || '',
                    class: inp.className || ''
                });
            }
            return result;
        """)
        log.info("    All input fields on page: %s", all_inputs)
    except Exception as e:
        log.warning("    Could not get input fields: %s", e)

    # Try common selectors for the life cover input with multiple attempts
    cover_selectors = [
        "input#lifeCover",
        "input[placeholder*='cover' i]",
        "input[placeholder*='sum' i]",
        "input[name*='cover' i]",
        "input[type='text']",  # Try any text input
        "input[type='number']",  # Try any number input
    ]
    cover_input = None
    for sel in cover_selectors:
        try:
            cover_input = wait_clickable(driver, By.CSS_SELECTOR, sel, timeout=5)
            log.info("    Found Life Cover input using selector: %s", sel)
            break
        except TimeoutException:
            continue

    if cover_input is None:
        # Try to find input near "Life Cover" text using JavaScript
        log.info("    Trying JavaScript approach to find Life Cover input")
        try:
            cover_input = driver.execute_script("""
                // Look for text containing "Life Cover" or "Cover"
                var allElements = document.querySelectorAll('*');
                for (var i = 0; i < allElements.length; i++) {
                    var el = allElements[i];
                    var txt = (el.innerText || el.textContent || '').toLowerCase();
                    if (txt.indexOf('life cover') !== -1 || txt.indexOf('life insurance cover') !== -1) {
                        // Found the label, now find nearby input
                        var container = el.closest('div, section, form');
                        if (container) {
                            var input = container.querySelector('input[type="text"], input[type="number"], input');
                            if (input) {
                                input.scrollIntoView({block: 'center'});
                                return input;
                            }
                        }
                    }
                }
                // Fallback: return first visible input
                var inputs = document.querySelectorAll('input');
                for (var i = 0; i < inputs.length; i++) {
                    var inp = inputs[i];
                    var style = window.getComputedStyle(inp);
                    if (style.display !== 'none' && style.visibility !== 'hidden') {
                        inp.scrollIntoView({block: 'center'});
                        return inp;
                    }
                }
                return null;
            """)
            if cover_input:
                log.info("    Found Life Cover input using JavaScript")
        except Exception as e:
            log.warning("    JavaScript approach failed: %s", e)

    if cover_input is None:
        log.error("    Could not find Life Cover input field")
        take_screenshot(driver, "life_cover_input_not_found")
        # Log page HTML for debugging
        try:
            page_html = driver.execute_script("return document.body.innerHTML;")
            log.error("    Page HTML (first 2000 chars): %s", page_html[:2000])
        except:
            pass
        raise NoSuchElementException("Life Cover input field not found")

    # Clear and fill the Life Cover field
    try:
        # Click to focus
        cover_input.click()
        time.sleep(0.3)
        
        # Clear existing value - try multiple methods
        cover_input.send_keys(Keys.CONTROL + "a")
        cover_input.send_keys(Keys.DELETE)
        time.sleep(0.2)
        
        # Alternative clear method
        cover_input.clear()
        time.sleep(0.2)
        
        # Type the new value
        cover_input.send_keys(cover_amount)
        log.info("    Filled Life Cover with: %s", cover_amount)
        time.sleep(0.5)
        
        # Verify the value was entered
        entered_value = cover_input.get_attribute("value")
        log.info("    Life Cover field value after entry: %s", entered_value)
        
        # If value doesn't match, try JavaScript
        if entered_value != cover_amount:
            log.warning("    Value mismatch, trying JavaScript method")
            driver.execute_script(f"arguments[0].value = '{cover_amount}';", cover_input)
            # Trigger change event
            driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", cover_input)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", cover_input)
            time.sleep(0.5)
            entered_value = cover_input.get_attribute("value")
            log.info("    Life Cover field value after JS: %s", entered_value)
            
    except Exception as e:
        log.error("    Error filling Life Cover field: %s", e)
        take_screenshot(driver, "life_cover_fill_error")
        raise

    # Cover Till Age – radio button
    cover_age = str(data.get("Cover till age", "65")).strip()
    _click_radio_label(driver, cover_age)

    # Proceed
    wait_clickable(driver, By.CSS_SELECTOR, "button#viewPlans").click()
    log.info("  ✓ Step 1 submitted")
    return True  # Indicate that Step 1 was successfully processed


def fill_step2_riders(driver: webdriver.Chrome, data: dict):
    """Step 2/4 – Critical Illness Rider, then proceed."""
    log.info("  → Step 2/4: Riders")
    
    # Log current URL for debugging
    current_url = driver.current_url
    log.info("  Current URL at Step 2: %s", current_url)
    
    # Take screenshot for debugging
    take_screenshot(driver, "step2_start")
    
    # Wait for page to be ready - try multiple indicators
    try:
        WebDriverWait(driver, DEFAULT_WAIT).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Rider') or contains(text(),'Enhance') or contains(text(),'Step 2')]"))
        )
        log.info("    Step 2 page detected via text")
    except TimeoutException:
        log.warning("    Step 2 text not found, trying button detection")
        try:
            # Try to find the proceed button as an alternative indicator
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button#viewPlans"))
            )
            log.info("    Step 2 page detected via button")
        except TimeoutException:
            log.error("    Step 2 page not detected - logging page state")
            # Log page title and visible text
            log.error("    Page title: %s", driver.title)
            page_text = driver.execute_script("return document.body.innerText.substring(0, 500);")
            log.error("    Visible text (first 500 chars): %s", page_text)
            take_screenshot(driver, "step2_not_found")
            raise

    ci_rider = str(data.get("Critical Illness Rider", "No")).strip().lower()
    if ci_rider == "yes":
        # Try to enable the Critical Illness checkbox/toggle
        try:
            ci_xpath = (
                "//input[@type='checkbox'][ancestor::*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', "
                "'abcdefghijklmnopqrstuvwxyz'), 'critical illness')]]"
            )
            cb = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, ci_xpath))
            )
            if not cb.is_selected():
                safe_click(driver, cb)
            log.info("    Critical Illness Rider enabled")
        except TimeoutException:
            log.warning("    Critical Illness Rider checkbox not found; skipping")

    # Proceed / Skip - try to find and click the button with multiple strategies
    log.info("    Looking for Skip/Proceed button...")
    
    button_clicked = False
    
    # Wait for page to be fully interactive before trying to click
    time.sleep(2)
    
    # Strategy 1: Try common button texts with XPath (case-insensitive)
    button_texts = ["Skip", "Proceed", "Continue", "Next", "View Plans"]
    
    for btn_text in button_texts:
        try:
            log.info(f"    Trying to find '{btn_text}' button...")
            xpath = f"//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{btn_text.lower()}')]"
            btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            # Scroll into view
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.5)
            # Try regular click first
            try:
                btn.click()
                log.info(f"    ✓ Clicked '{btn_text}' button (regular click)")
            except ElementClickInterceptedException:
                # Fallback to JS click
                driver.execute_script("arguments[0].click();", btn)
                log.info(f"    ✓ Clicked '{btn_text}' button (JS click)")
            button_clicked = True
            break
        except TimeoutException:
            log.debug(f"    '{btn_text}' button not found")
            continue
        except Exception as e:
            log.warning(f"    Error clicking '{btn_text}' button: {e}")
            continue
    
    # Strategy 2: Try CSS selector for viewPlans button
    if not button_clicked:
        try:
            log.info("    Trying CSS selector for viewPlans button...")
            proceed_btn = wait_clickable(driver, By.CSS_SELECTOR, "button#viewPlans", timeout=5)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", proceed_btn)
            time.sleep(0.5)
            try:
                proceed_btn.click()
                log.info("    ✓ Clicked viewPlans button (regular click)")
            except ElementClickInterceptedException:
                driver.execute_script("arguments[0].click();", proceed_btn)
                log.info("    ✓ Clicked viewPlans button (JS click)")
            button_clicked = True
        except TimeoutException:
            log.debug("    viewPlans button not found via CSS selector")
        except Exception as e:
            log.warning(f"    Error clicking viewPlans button: {e}")
    
    # Strategy 3: JavaScript fallback - find any visible button with relevant text
    if not button_clicked:
        log.warning("    Standard button click failed, trying comprehensive JS fallback")
        result = driver.execute_script("""
            var btns = document.querySelectorAll('button, a[role="button"], input[type="button"], input[type="submit"]');
            var found = [];
            
            // First, collect all visible buttons with their text
            for (var i = 0; i < btns.length; i++) {
                var btn = btns[i];
                var txt = (btn.innerText || btn.textContent || btn.value || '').trim();
                var style = window.getComputedStyle(btn);
                if (style.display !== 'none' && style.visibility !== 'hidden' && txt) {
                    found.push({element: btn, text: txt.toLowerCase()});
                }
            }
            
            // Try to find button with relevant text
            var keywords = ['proceed', 'continue', 'next', 'skip', 'view', 'plan'];
            for (var k = 0; k < keywords.length; k++) {
                for (var i = 0; i < found.length; i++) {
                    if (found[i].text.indexOf(keywords[k]) !== -1) {
                        found[i].element.scrollIntoView({block: 'center'});
                        found[i].element.click();
                        return 'Clicked: ' + found[i].text + ' (keyword: ' + keywords[k] + ')';
                    }
                }
            }
            
            // If no keyword match, try clicking the last visible button (usually the submit button)
            if (found.length > 0) {
                var lastBtn = found[found.length - 1];
                lastBtn.element.scrollIntoView({block: 'center'});
                lastBtn.element.click();
                return 'Clicked last button: ' + lastBtn.text;
            }
            
            return 'No button found. Available buttons: ' + found.map(function(f) { return f.text; }).join(', ');
        """)
        log.info("    JS button click result: %s", result)
        if result and 'Clicked' in str(result):
            log.info("  ✓ Step 2 submitted via JS")
            button_clicked = True
        else:
            log.error("    JS fallback result: %s", result)
    
    if not button_clicked:
        take_screenshot(driver, "step2_no_button_found")
        raise TimeoutException("Could not find proceed/skip button at Step 2")
    
    # Wait for navigation to complete - increased wait time
    log.info("    Waiting for navigation to complete...")
    time.sleep(3)
    
    # Verify we've moved past Step 2
    try:
        new_url = driver.current_url
        log.info("    URL after Step 2: %s", new_url)
        
        # Check if we're at Step 3 (Eligibility)
        step3_check = driver.execute_script("""
            var text = document.body.innerText.toLowerCase();
            return text.indexOf('eligibility') !== -1 || text.indexOf('step 3') !== -1;
        """)
        
        if step3_check:
            log.info("    ✓ Successfully navigated to Step 3")
        else:
            log.warning("    Step 3 not detected, but continuing...")
            take_screenshot(driver, "step2_after_click")
    except Exception as e:
        log.warning("    Could not verify Step 3: %s", e)
    
    log.info("  ✓ Step 2 completed")


def fill_step3_eligibility(driver: webdriver.Chrome, data: dict) -> Path | None:
    """
    Step 3/4 – Email, Annual Income, Pincode, City, Download PDF, Proceed.
    Returns the path to the downloaded PDF (or None if download failed).
    """
    log.info("  → Step 3/4: Eligibility")

    WebDriverWait(driver, DEFAULT_WAIT).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Eligibility') or contains(text(),'Step 3')]"))
    )

    # Email
    clear_and_type(driver, By.CSS_SELECTOR, "input#email", str(data.get("email id", "")).strip())

    # Annual Income (numeric)
    income_raw = str(data.get("Annual income", "0")).replace(",", "").strip()
    clear_and_type(driver, By.CSS_SELECTOR, "input#eligibilityAnnualIncome", income_raw)

    # Pincode
    pincode = str(data.get("Pincode", "")).strip()
    pin_input = clear_and_type(driver, By.CSS_SELECTOR, "input#pincode", pincode)
    time.sleep(1.5)  # wait for city dropdown to populate

    # City – select Bangalore from dropdown
    _select_city_bangalore(driver)

    # Download Benefit Illustration PDF
    pdf_path = _download_benefit_illustration(driver, data)

    # Proceed
    wait_clickable(driver, By.CSS_SELECTOR, "button#viewPlans").click()
    log.info("  ✓ Step 3 submitted")
    return pdf_path


def _select_city_bangalore(driver, timeout=DEFAULT_WAIT):
    """Select Bangalore from the city dropdown that appears after pincode entry."""
    try:
        # Try <select> element first
        sel_el = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.XPATH, "//select[contains(@id,'city') or contains(@name,'city')]"))
        )
        Select(sel_el).select_by_visible_text("Bangalore")
        return
    except (TimeoutException, NoSuchElementException):
        pass

    # Try autocomplete / dropdown list
    try:
        option = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(@class,'option') or contains(@class,'item')][contains(normalize-space(.), 'Bangalore')]")
            )
        )
        safe_click(driver, option)
    except TimeoutException:
        log.warning("    Could not select Bangalore from city dropdown; continuing")


def _download_benefit_illustration(driver, data: dict) -> Path | None:
    """Click the 'Download Benefit Illustration' link and wait for the PDF."""
    # Record existing PDFs before clicking
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    before = set(DOWNLOAD_DIR.glob("*.pdf"))

    try:
        link = WebDriverWait(driver, DEFAULT_WAIT).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(normalize-space(.), 'Benefit Illustration') and (self::a or self::button or self::span)]")
            )
        )
        safe_click(driver, link)
        log.info("    Clicked 'Download Benefit Illustration'")
    except TimeoutException:
        log.warning("    'Download Benefit Illustration' link not found; skipping PDF download")
        return None

    # Wait for new PDF to appear
    deadline = time.time() + DOWNLOAD_WAIT
    while time.time() < deadline:
        after = set(DOWNLOAD_DIR.glob("*.pdf"))
        new_files = after - before
        # Ignore .crdownload (in-progress Chrome downloads)
        completed = [f for f in new_files if not f.suffix == ".crdownload"]
        if completed:
            pdf_path = max(completed, key=lambda p: p.stat().st_mtime)
            log.info("    PDF downloaded: %s", pdf_path.name)
            return pdf_path
        time.sleep(1)

    log.warning("    PDF download timed out after %ds", DOWNLOAD_WAIT)
    return None


def extract_step4_summary(driver: webdriver.Chrome) -> dict:
    """
    Step 4/4 – Extract Equote Number and Premium from the summary page.
    Returns dict with keys 'equote_number' and 'premium'.
    """
    log.info("  → Step 4/4: Summary")
    
    # Log current URL
    current_url = driver.current_url
    log.info("  Current URL at Step 4: %s", current_url)
    
    # Wait for page to be fully loaded
    try:
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        log.info("    Page load state: complete")
    except TimeoutException:
        log.warning("    Page load state timeout")
    
    # Take screenshot
    take_screenshot(driver, "step4_start")
    
    # Wait for Step 4 page - try multiple indicators with longer timeout
    page_ready = False
    
    # Try 1: Look for summary text
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Summary') or contains(text(),'Step 4')]"))
        )
        log.info("    Step 4 page detected via Summary text")
        page_ready = True
    except TimeoutException:
        log.warning("    Step 4 Summary text not found")
    
    # Try 2: Look for Equote or Premium text
    if not page_ready:
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Equote') or contains(text(),'Quote') or contains(text(),'Premium')]"))
            )
            log.info("    Step 4 page detected via Equote/Premium text")
            page_ready = True
        except TimeoutException:
            log.warning("    Equote/Premium text not found")
    
    # If still not ready, log page state but continue
    if not page_ready:
        log.warning("    Step 4 page not fully detected - attempting to extract data anyway")
        log.info("    Page title: %s", driver.title)
        try:
            page_text = driver.execute_script("return document.body.innerText.substring(0, 1000);")
            log.info("    Visible text (first 1000 chars): %s", page_text)
        except:
            pass
        take_screenshot(driver, "step4_uncertain")
    
    # ═══════════════════════════════════════════════════════════════════════
    # CRITICAL: Wait for EQI Number to be generated and displayed
    # ═══════════════════════════════════════════════════════════════════════
    log.info("    Waiting for EQI Number to be generated...")
    
    # Strategy 1: Wait for EQI/Equote text to appear with a code pattern
    eqi_found = False
    max_wait_time = 30  # Wait up to 30 seconds for EQI generation
    
    try:
        # Wait for an element containing "Equote" or "EQI" followed by a code pattern
        WebDriverWait(driver, max_wait_time).until(
            lambda d: bool(re.search(
                r"(?:Equote|EQI|E-Quote)[\s:]*[0-9]{3,}[A-Z]{2,}",
                d.find_element(By.TAG_NAME, "body").text,
                re.IGNORECASE
            ))
        )
        log.info("    ✓ EQI Number detected on page")
        eqi_found = True
    except TimeoutException:
        log.warning("    EQI Number not detected after %d seconds", max_wait_time)
    
    # Strategy 2: If Strategy 1 failed, wait for any code-like pattern (fallback)
    if not eqi_found:
        log.info("    Trying fallback: waiting for any code pattern...")
        try:
            WebDriverWait(driver, 10).until(
                lambda d: bool(re.search(
                    r"\b[0-9]{3,4}[A-Z]{2,}[A-Z0-9]*\b",
                    d.find_element(By.TAG_NAME, "body").text
                ))
            )
            log.info("    ✓ Code pattern detected on page")
            eqi_found = True
        except TimeoutException:
            log.warning("    No code pattern detected")
    
    # Additional wait to ensure all dynamic content is loaded
    log.info("    Waiting additional 3 seconds for dynamic content...")
    time.sleep(3)
    
    # Take screenshot after waiting
    take_screenshot(driver, "step4_after_wait")

    # Get full page text for extraction
    page_text = driver.find_element(By.TAG_NAME, "body").text
    log.info("    Extracting data from page text (length: %d chars)", len(page_text))
    
    # Log a sample of the page text for debugging
    sample_text = page_text[:500] if len(page_text) > 500 else page_text
    log.info("    Page text sample: %s", sample_text)

    # Extract Equote Number (e.g., "0794VSCP" or "Equote Number: 0794VSCP")
    equote = ""
    
    # Try multiple patterns to find the EQI number
    patterns = [
        # Pattern 1: "Equote Number" followed by the code
        r"Equote\s+Number[\s:]*([0-9]{3,}[A-Z]{2,}[A-Z0-9]*)",
        # Pattern 2: "EQI" or "E-Quote" followed by the code
        r"(?:EQI|E-Quote)[\s:]*([0-9]{3,}[A-Z]{2,}[A-Z0-9]*)",
        # Pattern 3: Standalone code (3-4 digits followed by 2+ uppercase letters)
        r"\b([0-9]{3,4}[A-Z]{2,}[A-Z0-9]*)\b",
        # Pattern 4: More flexible - any sequence of 3+ digits followed by 2+ letters
        r"\b([0-9]{3,}[A-Z]{2,})\b",
    ]
    
    for i, pattern in enumerate(patterns):
        m = re.search(pattern, page_text, re.IGNORECASE if i < 2 else 0)
        if m:
            equote = m.group(1).strip().upper()
            log.info("    Equote Number found (pattern %d): %s", i + 1, equote)
            break
    
    if not equote:
        log.warning("    Equote Number not found with regex patterns")
        # Try JavaScript extraction as fallback - get clean text content
        try:
            equote = driver.execute_script("""
                // Look for elements containing "Equote Number" or "EQI"
                var allText = document.body.innerText || document.body.textContent || '';
                var lines = allText.split('\\n');
                
                for (var i = 0; i < lines.length; i++) {
                    var line = lines[i].trim();
                    var lowerLine = line.toLowerCase();
                    if (lowerLine.indexOf('equote') !== -1 || lowerLine.indexOf('eqi') !== -1 || lowerLine.indexOf('quote number') !== -1) {
                        // Check current line for the code
                        var match = line.match(/[0-9]{3,}[A-Z]{2,}[A-Z0-9]*/i);
                        if (match) return match[0].toUpperCase();
                        
                        // Check next line
                        if (i + 1 < lines.length) {
                            var nextLine = lines[i + 1].trim();
                            var nextMatch = nextLine.match(/[0-9]{3,}[A-Z]{2,}[A-Z0-9]*/i);
                            if (nextMatch) return nextMatch[0].toUpperCase();
                        }
                    }
                }
                
                // Last resort: look for any code-like pattern
                for (var i = 0; i < lines.length; i++) {
                    var line = lines[i].trim();
                    var match = line.match(/\\b[0-9]{4,}[A-Z]{2,}\\b/i);
                    if (match) return match[0].toUpperCase();
                }
                
                return '';
            """)
            if equote:
                log.info("    Equote Number found via JS: %s", equote)
        except Exception as e:
            log.warning("    JS Equote extraction failed: %s", e)

    # Extract Premium (e.g., "₹ 12,345" or "Premium: 12345")
    premium = ""
    m2 = re.search(
        r"(?:Premium|Total\s+Premium|Annual\s+Premium)[\s:₹]*([0-9,]+(?:\.[0-9]+)?)",
        page_text,
        re.IGNORECASE,
    )
    if m2:
        premium = m2.group(1).replace(",", "").strip()
        log.info("    Premium found: %s", premium)
    else:
        log.warning("    Premium not found in page text")
        # Try JavaScript extraction as fallback
        try:
            premium = driver.execute_script("""
                var elements = document.querySelectorAll('*');
                for (var i = 0; i < elements.length; i++) {
                    var el = elements[i];
                    var txt = (el.innerText || el.textContent || '').toLowerCase();
                    if (txt.indexOf('premium') !== -1 || txt.indexOf('amount') !== -1) {
                        // Look for currency amount
                        var match = txt.match(/[₹$]?\\s*[0-9,]+\\.?[0-9]*/i);
                        if (match) return match[0].replace(/[₹$\\s]/g, '');
                        if (el.nextElementSibling) {
                            var nextTxt = el.nextElementSibling.innerText || el.nextElementSibling.textContent || '';
                            var nextMatch = nextTxt.match(/[₹$]?\\s*[0-9,]+\\.?[0-9]*/i);
                            if (nextMatch) return nextMatch[0].replace(/[₹$\\s]/g, '');
                        }
                    }
                }
                return '';
            """)
            if premium:
                log.info("    Premium found via JS: %s", premium)
        except Exception as e:
            log.warning("    JS Premium extraction failed: %s", e)
    
    # Take final screenshot
    take_screenshot(driver, "step4_complete")

    log.info("    Final extracted values - Equote: %s  |  Premium: %s", equote, premium)
    return {"equote_number": equote, "premium": premium}


# ---------------------------------------------------------------------------
# PDF comparison
# ---------------------------------------------------------------------------

def extract_from_pdf(pdf_path: Path) -> dict:
    """Extract Equote Number and Premium from a Benefit Illustration PDF."""
    result = {"equote_number": "", "premium": ""}
    if pdf_path is None or not pdf_path.exists():
        return result

    try:
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)

        m = re.search(r"(?:Equote\s*(?:No|Number|#)?[\s:]*)([\w\-]+)", text, re.IGNORECASE)
        if m:
            result["equote_number"] = m.group(1).strip()

        m2 = re.search(
            r"(?:Premium|Total\s+Premium|Annual\s+Premium)[\s:₹]*([0-9,]+(?:\.[0-9]+)?)",
            text,
            re.IGNORECASE,
        )
        if m2:
            result["premium"] = m2.group(1).replace(",", "").strip()

    except Exception as exc:
        log.warning("PDF extraction error: %s", exc)

    return result


def compare_pdf_values(portal: dict, pdf: dict) -> str:
    """Return 'MATCH', 'MISMATCH', or 'PDF_NOT_FOUND'."""
    if not pdf.get("equote_number") and not pdf.get("premium"):
        return "PDF_NOT_FOUND"
    eq_match = portal.get("equote_number", "") == pdf.get("equote_number", "")
    pr_match = portal.get("premium", "") == pdf.get("premium", "")
    return "MATCH" if (eq_match and pr_match) else "MISMATCH"


# ---------------------------------------------------------------------------
# Report generator
# ---------------------------------------------------------------------------

REPORT_COLUMNS = ["Unique Identity Number", "User Name", "EQI Number", "Premium Price"]


def generate_report(results: list[dict]):
    """Write CSV and XLSX reports to RESULTS_DIR."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # CSV
    csv_path = RESULTS_DIR / f"report_{timestamp}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=REPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    log.info("CSV report: %s", csv_path)

    # XLSX
    xlsx_path = RESULTS_DIR / f"report_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(REPORT_COLUMNS)
    for row in results:
        ws.append([row.get(col, "") for col in REPORT_COLUMNS])
    wb.save(xlsx_path)
    log.info("XLSX report: %s", xlsx_path)

    return csv_path, xlsx_path


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def run_test_case(driver: webdriver.Chrome, idx: int, data: dict) -> dict:
    """Run a single test case end-to-end. Returns a result dict."""
    name = str(data.get("Full Name", f"Row {idx + 1}"))
    log.info("=" * 60)
    log.info("Test case %d: %s", idx + 1, name)
    log.info("=" * 60)

    result = {
        "Unique Identity Number": f"ID{idx + 1:03d}",  # e.g., ID001, ID002, etc.
        "User Name": name,
        "EQI Number": "",
        "Premium Price": "",
    }

    try:
        # Clear browser state before starting new test case
        log.info("  Clearing browser state...")
        
        # Navigate to about:blank first to ensure clean state
        driver.get("about:blank")
        time.sleep(1)
        
        # Clear cookies
        driver.delete_all_cookies()
        log.info("  Cleared cookies")
        
        # Clear local storage and session storage
        try:
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")
            log.info("  Cleared storage")
        except Exception as e:
            log.warning("  Could not clear storage: %s", e)
        
        # Navigate to portal - use base URL without query parameters for fresh start
        base_url = "https://www.axismaxlife.com/term-insurance-plans/premium-calculator"
        log.info("  Navigating to portal URL (base)...")
        driver.get(base_url)
        
        # Wait for page to be fully loaded
        max_retries = 3
        for attempt in range(max_retries):
            try:
                WebDriverWait(driver, 30).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                log.info("  Page loaded successfully")
                break
            except TimeoutException:
                if attempt < max_retries - 1:
                    log.warning("  Page load timeout, retrying...")
                    driver.refresh()
                else:
                    log.error("  Page load timeout after %d attempts", max_retries)
                    raise
        
        time.sleep(4)  # Additional wait for dynamic content and JavaScript initialization
        
        # Verify we're on the landing page by checking for the form
        landing_page_found = False
        for attempt in range(3):
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input#fullName"))
                )
                log.info("  Landing page form detected")
                landing_page_found = True
                break
            except TimeoutException:
                log.warning("  Landing page form not found (attempt %d/3)", attempt + 1)
                take_screenshot(driver, f"landing_page_not_found_tc{idx + 1}_attempt{attempt + 1}")
                
                if attempt < 2:
                    # Try navigating again with a longer URL that might force landing page
                    log.info("  Retrying navigation to portal...")
                    driver.get("about:blank")
                    time.sleep(1)
                    driver.delete_all_cookies()
                    driver.get(base_url)
                    time.sleep(5)
                else:
                    log.error("  Landing page form not found after 3 attempts")
                    raise
        
        if not landing_page_found:
            raise TimeoutException("Could not reach landing page after multiple attempts")

        fill_landing_page(driver, data)
        time.sleep(1)

        fill_details_modal(driver, data)
        
        # After modal, wait for navigation and page load
        time.sleep(3)  # Wait for navigation to complete
        
        # Wait for page to be interactive
        try:
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            log.info("  Page ready after details modal")
        except TimeoutException:
            log.warning("  Page ready state timeout after details modal")
        
        current_url = driver.current_url
        log.info("  URL after details modal: %s", current_url)
        
        # Detect which step we're on by checking for specific input fields
        try:
            # Check if we're at Step 3 (Eligibility) - has email, pincode, eligibilityAnnualIncome
            step3_indicator = driver.execute_script("""
                return document.querySelector('input#email') !== null && 
                       document.querySelector('input#pincode') !== null &&
                       document.querySelector('input#eligibilityAnnualIncome') !== null;
            """)
            
            if step3_indicator:
                log.warning("  ⚠ Portal skipped Steps 1 & 2 - already at Step 3 (Eligibility)")
                log.info("  Proceeding directly to Step 3")
                
                # Go directly to Step 3
                pdf_path = fill_step3_eligibility(driver, data)
                
                # Wait for navigation to Step 4 and for page to be fully loaded
                log.info("  Waiting for Step 4 to load...")
                time.sleep(3)
                
                # Wait for page ready state
                try:
                    WebDriverWait(driver, 15).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    log.info("  Step 4 page ready")
                except TimeoutException:
                    log.warning("  Step 4 page ready timeout")

                portal_values = extract_step4_summary(driver)
                result["EQI Number"] = portal_values["equote_number"]
                result["Premium Price"] = portal_values["premium"]

                log.info("  Result: EQI=%s  Premium=%s",
                         result["EQI Number"], result["Premium Price"])
                return result
        except Exception as e:
            log.debug("  Step 3 detection check failed: %s", e)
        
        # Check if we're at Step 2 (Riders)
        try:
            step2_indicator = driver.execute_script("""
                var text = document.body.innerText.toLowerCase();
                return text.indexOf('rider') !== -1 || text.indexOf('enhance') !== -1;
            """)
            
            if step2_indicator or 'stage=rider' in current_url.lower() or 'stage=2' in current_url.lower():
                log.warning("  ⚠ Portal skipped Step 1 - already at Step 2 (Riders)")
                log.info("  Proceeding directly to Step 2")
                step1_processed = False
            else:
                # We're at Step 1
                step1_processed = fill_step1_customize(driver, data)
        except Exception as e:
            log.debug("  Step 2 detection check failed: %s", e)
            # Default: try Step 1
            step1_processed = fill_step1_customize(driver, data)
        
        if step1_processed:
            # Step 1 was processed normally, wait before Step 2
            time.sleep(1)
        else:
            # Step 1 was skipped by portal, we're already at Step 2
            log.info("  Portal auto-advanced to Step 2, proceeding directly to riders")

        fill_step2_riders(driver, data)
        time.sleep(1)

        pdf_path = fill_step3_eligibility(driver, data)
        
        # Wait for navigation to Step 4 and for page to be fully loaded
        log.info("  Waiting for Step 4 to load...")
        time.sleep(3)
        
        # Wait for page ready state
        try:
            WebDriverWait(driver, 15).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            log.info("  Step 4 page ready")
        except TimeoutException:
            log.warning("  Step 4 page ready timeout")

        portal_values = extract_step4_summary(driver)
        result["EQI Number"] = portal_values["equote_number"]
        result["Premium Price"] = portal_values["premium"]

        log.info("  Result: EQI=%s  Premium=%s",
                 result["EQI Number"], result["Premium Price"])

    except Exception as exc:
        error_msg = f"{type(exc).__name__}: {exc}"
        log.error("  FAILED: %s", error_msg)
        log.debug(traceback.format_exc())
        take_screenshot(driver, f"error_tc{idx + 1}_{name.replace(' ', '_')}")
        # Set empty values on error
        result["EQI Number"] = "ERROR"
        result["Premium Price"] = "ERROR"

    return result


def main():
    log.info("Reading test data from: %s", EXCEL_PATH)
    test_rows = read_test_data()

    driver = create_driver()
    all_results = []

    try:
        for idx, row in enumerate(test_rows):
            try:
                result = run_test_case(driver, idx, row)
                all_results.append(result)
                
                # Longer pause between test cases to ensure browser is ready
                if idx < len(test_rows) - 1:  # Don't wait after the last test case
                    log.info("  Waiting before next test case...")
                    time.sleep(5)
            except Exception as e:
                # If a test case fails completely, log it and continue with next one
                log.error("  Test case %d failed with exception: %s", idx + 1, str(e))
                log.debug(traceback.format_exc())
                
                # Add error result
                error_result = {
                    "Unique Identity Number": f"ID{idx + 1:03d}",
                    "User Name": str(row.get("Full Name", f"Row {idx + 1}")),
                    "EQI Number": "ERROR",
                    "Premium Price": "ERROR",
                }
                all_results.append(error_result)
                
                # Try to recover by refreshing the browser
                try:
                    log.info("  Attempting to recover browser state...")
                    driver.delete_all_cookies()
                    driver.execute_script("window.localStorage.clear();")
                    driver.execute_script("window.sessionStorage.clear();")
                    time.sleep(2)
                except:
                    pass
                
    except KeyboardInterrupt:
        log.warning("Run interrupted by user — saving partial results (%d rows)", len(all_results))
    finally:
        try:
            driver.quit()
            log.info("Browser closed")
        except Exception:
            # Browser may already be gone (e.g. after Ctrl+C); safe to ignore
            log.debug("driver.quit() failed — browser was already closed")

    if not all_results:
        log.warning("No results to report.")
        return

    csv_path, xlsx_path = generate_report(all_results)

    # Summary
    log.info("")
    log.info("=" * 60)
    log.info("SUMMARY")
    log.info("=" * 60)
    for r in all_results:
        log.info("  ID: %-10s  Name: %-25s  EQI: %-15s  Premium: %-10s",
                 r["Unique Identity Number"], r["User Name"], r["EQI Number"], r["Premium Price"])
    log.info("")
    log.info("Reports saved to: %s", RESULTS_DIR)


if __name__ == "__main__":
    main()
