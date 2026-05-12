import openpyxl

wb = openpyxl.load_workbook(r'test data\insurance_test_data.xlsx')
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print('Headers:', headers)
print()

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=7, values_only=True)):
    data = dict(zip(headers, row))
    print(f'Row {i+1}:')
    print(f'  Full Name: {data.get("Full Name")}')
    print(f'  Life cover: {data.get("Life cover")}')
    print(f'  Cover till age: {data.get("Cover till age")}')
    print()
